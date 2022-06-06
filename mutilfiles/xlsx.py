import xlrd 
from openpyxl import load_workbook
wb = xlrd.open_workbook('./data.xlsx')
wb1 = load_workbook('./data.xlsx')
sheets_names = wb1.sheetnames
ws2= wb1[sheets_names[0]]
print("open ok")
gcc_commond = open('gcc_commond.txt','r')
list1 = []
for line in gcc_commond.readlines():
	line = line.strip()
	list1.append(line)

#获取成绩
score=0
gcc_commond = open('grade1.txt','r')
for line in gcc_commond.readlines():
	line = line.strip()
	score = int(line)


#count_record 和 name_record 分别代表 次数 和 姓名
name_record = list1[1]
count_record = int(list1[2]) 
print(name_record)
sht = wb.sheets()[0]

num_count = 0
name_col = sht.col_values(0,start_rowx=0,end_rowx=None)
for x in name_col:
	num_count += 1
	if(str(x) == str(name_record)):
		print("find it!!!")
		print(x)
		break
#ws2.cell(5,4).value = '3'

'''
for x in range(2,31):
	cell = 'A' + str(x)
#	print(sht.range(cell).value)
	if (str(sht.range(cell).value) == str(name_record)):
		num_count = x
		print("find it!!!")
		print(num_count)
		break
'''

count_record += 65 #转换成 ascill
tar = chr(count_record)
tar = tar + str(num_count)
print(tar)
ws2[tar] = str(score)
print(ws2['F13'].value)
